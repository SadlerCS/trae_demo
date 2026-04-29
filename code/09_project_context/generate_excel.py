import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_finance_excel():
    file_path = r'c:\AAAAAAAAAAAAAAAAA\workplace\ai_workplace\trae_workplace\trae_demo\code\09_project_context\Finance_Tracking_2026_04.xlsx'
    wb = Workbook()
    
    # --- 1. 支出工作表 ---
    ws_expense = wb.active
    ws_expense.title = "支出明细"
    
    expense_headers = ["日期", "项目说明", "金额 (USD)"]
    expense_data = [
        ["2026-04-20", "谷歌号美区", 15],
        ["2026-04-20", "号1-pixel验证换成品号美区", 15.45],
        ["2026-04-22", "号2成品号美区30质保", 30],
        ["2026-04-23", "谷歌号美区", 7.26],
        ["2026-04-23", "号3-成品号美区30质保", 30],
        ["2026-04-24", "号4-成品号美区", 30.9],
        ["2026-04-24", "号5-成品号美区30质保", 33],
        ["2026-04-24", "号6-成品号美区", 23],
        ["2026-04-27", "谷歌号美区", 9],
        ["2026-04-27", "谷歌号美区", 9],
        ["2026-04-27", "谷歌号美区 (白嫖)", 0],
        ["2026-04-28", "谷歌号美区", 9.79],
    ]
    
    ws_expense.append(expense_headers)
    for row in expense_data:
        ws_expense.append(row)
    
    # 添加总计行
    last_row = len(expense_data) + 1
    ws_expense.append(["", "总计", f"=SUM(C2:C{last_row})"])
    
    # --- 2. 收入工作表 ---
    ws_income = wb.create_sheet("收入明细")
    income_headers = ["日期", "项目说明", "原始金额", "税率 (1.6%)", "税后收入", "备注"]
    # 备注：闲鱼币 1000:1
    income_data = [
        ["2026-04-22", "号6+2美区谷歌", 180, 0.016, "=C2*(1-D2)", ""],
        ["2026-04-23", "号5", 125, 0.016, "=C3*(1-D3)", ""],
        ["2026-04-24", "号5", 125, 0.016, "=C4*(1-D4)", ""],
        ["2026-04-25", "号3", 75, 0.016, "=C5*(1-D5)", ""],
        ["2026-04-27", "号5+号1+3美区谷歌", 187.5, 0.016, "=C6*(1-D6)", "含闲鱼币2500"],
        ["2026-04-29", "号1+1美区谷歌", 54.9, 0.016, "=C7*(1-D7)", "39.9+15, 含闲鱼币3990"],
    ]
    
    ws_income.append(income_headers)
    for row in income_data:
        ws_income.append(row)
        
    last_row_income = len(income_data) + 1
    ws_income.append(["", "总计", f"=SUM(C2:C{last_row_income})", "", f"=SUM(E2:E{last_row_income})", ""])

    # --- 3. 汇总工作表 ---
    ws_summary = wb.create_sheet("月度汇总", 0)
    ws_summary.append(["类目", "金额"])
    ws_summary.append(["总收入 (税后)", f"='收入明细'!E{last_row_income+1}"])
    ws_summary.append(["总支出", f"='支出明细'!C{last_row+1}"])
    ws_summary.append(["净利润", "=B2-B3"])
    
    # --- 样式美化 ---
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center")
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
        
        # 自动调整列宽
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    print(f"Excel 文件已生成: {file_path}")

if __name__ == "__main__":
    create_finance_excel()
