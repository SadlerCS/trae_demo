import pandas as pd
import os
import re
import argparse
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font

def sanitize_filename(filename):
    """清理文件名中的非法字符"""
    if not filename or pd.isna(filename):
        return "未分类"
    # 替换 Windows/Linux 下的非法字符
    return re.sub(r'[\\/*?:"<>|]', '_', str(filename))

def find_category_column(df):
    """自动识别分类列名"""
    possible_names = ['商品分类', '分类', '类别', 'category', 'type', 'class']
    for col in df.columns:
        if str(col).lower() in possible_names:
            return col
    return None

def add_summary_row(writer, df, sheet_name, output_format):
    """为 Excel 文件添加汇总行"""
    if output_format.lower() != 'xlsx':
        return
    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    last_row = len(df) + 1 # +1 for header
    
    # 汇总数据
    count = len(df)
    total_sales = 0
    avg_price = 0
    
    # 尝试找到销售额和单价列
    sales_col = None
    price_col = None
    for i, col in enumerate(df.columns):
        if '销售额' in str(col) or 'amount' in str(col).lower():
            sales_col = i
        if '单价' in str(col) or 'price' in str(col).lower():
            price_col = i
            
    # 计算汇总值
    if sales_col is not None:
        total_sales = df.iloc[:, sales_col].sum()
    if price_col is not None:
        avg_price = df.iloc[:, price_col].mean()

    # 写入汇总行
    bold_font = Font(bold=True)
    summary_label_cell = worksheet.cell(row=last_row + 2, column=1, value="汇总信息")
    summary_label_cell.font = bold_font
    
    worksheet.cell(row=last_row + 3, column=1, value=f"记录条数: {count}").font = bold_font
    if sales_col is not None:
        worksheet.cell(row=last_row + 4, column=1, value=f"销售额合计: {total_sales:.2f}").font = bold_font
    if price_col is not None:
        worksheet.cell(row=last_row + 5, column=1, value=f"平均单价: {avg_price:.2f}").font = bold_font

def process_sales_data(input_file, output_dir, output_format='xlsx', include_summary=True):
    """核心处理函数"""
    start_time = datetime.now()
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 设置日志
    log_file = os.path.join(output_dir, f"split_log_{datetime.now().strftime('%Y%m%d')}.txt")
    logging.basicConfig(filename=log_file, level=logging.INFO, 
                        format='%(asctime)s - %(message)s', encoding='utf-8')
    
    logger = logging.getLogger()
    # 同时也输出到控制台
    console = logging.StreamHandler()
    logger.addHandler(console)

    logger.info(f"开始处理文件: {input_file}")
    
    try:
        # 读取所有工作表
        all_sheets = pd.read_excel(input_file, sheet_name=None)
        all_data = pd.concat(all_sheets.values(), ignore_index=True)
        
        total_records = len(all_data)
        logger.info(f"读取到总记录数: {total_records}")

        # 识别分类列
        cat_col = find_category_column(all_data)
        if not cat_col:
            logger.error("未找到‘商品分类’列，请检查文件结构。")
            # 如果没找到，默认将所有数据归为“未分类”
            all_data['temp_category'] = '未分类'
            cat_col = 'temp_category'
        
        # 按分类分组
        grouped = all_data.groupby(cat_col, dropna=False)
        
        file_count = 0
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        for cat_name, group in grouped:
            # 清理分类名称
            safe_cat_name = sanitize_filename(cat_name)
            filename = f"{safe_cat_name}_{timestamp}.{output_format}"
            file_path = os.path.join(output_dir, filename)
            
            if output_format == 'xlsx':
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    group.to_excel(writer, index=False, sheet_name='Sheet1')
                    if include_summary:
                        add_summary_row(writer, group, 'Sheet1', output_format)
            else:
                group.to_csv(file_path, index=False, encoding='utf-8-sig')
                if include_summary:
                    # CSV 简单添加汇总行
                    with open(file_path, 'a', encoding='utf-8-sig') as f:
                        f.write("\n汇总信息\n")
                        f.write(f"记录条数,{len(group)}\n")
                        sales_col = [c for c in group.columns if '销售额' in str(c) or 'amount' in str(c).lower()]
                        if sales_col:
                            f.write(f"销售额合计,{group[sales_col[0]].sum():.2f}\n")

            logger.info(f"生成文件: {filename}, 记录数: {len(group)}")
            file_count += 1

        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        logger.info(f"处理完成！生成文件总数: {file_count}, 总耗时: {duration:.2f}秒")
        
    except Exception as e:
        logger.error(f"处理过程中发生错误: {str(e)}")
        raise

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="销售数据分类拆分工具")
    parser.add_argument("--input", default="sales_data.xlsx", help="输入文件路径")
    parser.add_argument("--output_dir", default="./classified_output/", help="输出目录")
    parser.add_argument("--format", choices=['xlsx', 'csv'], default="xlsx", help="输出格式")
    parser.add_argument("--no_summary", action="store_true", help="不生成汇总行")
    
    args = parser.parse_args()
    
    process_sales_data(
        input_file=args.input,
        output_dir=args.output_dir,
        output_format=args.format,
        include_summary=not args.no_summary
    )
